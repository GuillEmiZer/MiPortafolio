import openpyxl
from tabulate import tabulate

wb = openpyxl.load_workbook("Libro1.xlsx")
ws = wb.active

numero_base = 0

if ws['B1'].value is None:
    numero_base = int(input("INGRESE NÚMERO VENTA BASE: "))
    ws.cell(row=1, column=2).value = numero_base
else:
    numero_base = ws.cell(row=1, column=2).value

SALTO = 26
ventas_ingresadas = 0
productos_ingresados = 0
pagos_ingresados = 0
A = 1
i = 0
fila_num_venta = 3
columna2 = 2
columna3 = 3
columna4 = 4
columna5 = 5
columna6 = 6
columna7 = 7
columna8 = 8
columna9 = 9
fila_fecha = 4
fila_plataforma = 5
fila_nombre_comprador = 6
fila_usuario_comprador = 7
fila_condicion_iva = 8
fila_cuit_dni = 9
fila_jurisdiccion = 10
fila_producto = 11
fila_envio = 15
fila16 = 16
fila17 = 17
fila23 = 23

while i < A:
    pregunta = input("¿Desea ingresar otra venta? SI/NO: ")
    if pregunta.upper() == "SI":
        for fila in range(fila_num_venta, ws.max_row + 1, SALTO):
            if ws.cell(row=fila, column=columna2).value is not None:
                ventas_ingresadas += 1

        num_venta = numero_base + (ventas_ingresadas + 1)
        print(f"VENTA N°{num_venta}")
        ws.cell(row=fila_num_venta + (SALTO*ventas_ingresadas), column=columna2).value = num_venta
        fecha = input("Fecha de Venta (00/00/00): ")
        ws.cell(row=fila_fecha+(SALTO*ventas_ingresadas), column=columna2).value = fecha
        plataforma = input("Plataforma de Venta: ")
        ws.cell(row=fila_plataforma + (SALTO * ventas_ingresadas), column=columna2).value = plataforma
        nombre_comprador = input("Nombre del Comprador: ")
        ws.cell(row=fila_nombre_comprador + (SALTO * ventas_ingresadas), column=columna2).value = nombre_comprador
        usuario_comprador = input("Usuario del Comprador: ")
        ws.cell(row=fila_usuario_comprador + (SALTO * ventas_ingresadas), column=columna2).value = usuario_comprador
        condicion_iva = input("Condicion Frente a IVA: ")
        ws.cell(row=fila_condicion_iva + (SALTO * ventas_ingresadas), column=columna2).value = condicion_iva
        cuit_dni = input("CUIT/DNI Facturación: ")
        ws.cell(row=fila_cuit_dni + (SALTO * ventas_ingresadas), column=columna2).value = cuit_dni
        jurisdiccion = input("Jurisdicción: ")
        ws.cell(row=fila_jurisdiccion + (SALTO * ventas_ingresadas), column=columna2).value = jurisdiccion
        print("NO SE INGRESARON PRODUCTOS AL PEDIDO")

        LIMITE_PRODUCTO = 5
        while productos_ingresados <= LIMITE_PRODUCTO:
            pregunta = input("¿Ingresar Nuevo Producto al Pedido? SI/NO: ")
            if pregunta.upper() == "SI":
                desc_prod = input("Descripción del Producto: ")
                ws.cell(row=fila_producto + productos_ingresados + (SALTO*ventas_ingresadas), column=columna4).value = desc_prod
                precio_unit = int(input("Precio Unitario: "))
                ws.cell(row=fila_producto + productos_ingresados + (SALTO*ventas_ingresadas), column=columna6).value = precio_unit
                cantidad_prod = int(input("Cantidad (unidades): "))
                ws.cell(row=fila_producto + productos_ingresados + (SALTO*ventas_ingresadas), column=columna7).value = cantidad_prod

                data = []
                if ws.cell(row=fila_producto + (SALTO * ventas_ingresadas), column=columna4).value is not None:
                    data.append([
                        ws.cell(row=fila_producto + (SALTO * ventas_ingresadas), column=columna4).value,
                        ws.cell(row=fila_producto + (SALTO * ventas_ingresadas), column=columna6).value,
                        ws.cell(row=fila_producto + (SALTO * ventas_ingresadas), column=columna7).value
                    ])
                if ws.cell(row=fila_producto + 1 + (SALTO * ventas_ingresadas), column=columna4).value is not None:
                    data.append([
                        ws.cell(row=fila_producto + 1 + (SALTO * ventas_ingresadas), column=columna4).value,
                        ws.cell(row=fila_producto + 1 + (SALTO * ventas_ingresadas), column=columna6).value,
                        ws.cell(row=fila_producto + 1 + (SALTO * ventas_ingresadas), column=columna7).value
                    ])
                if ws.cell(row=fila_producto + 2 + (SALTO * ventas_ingresadas), column=columna4).value is not None:
                    data.append([
                        ws.cell(row=fila_producto + 2 + (SALTO * ventas_ingresadas), column=columna4).value,
                        ws.cell(row=fila_producto + 2 + (SALTO * ventas_ingresadas), column=columna6).value,
                        ws.cell(row=fila_producto + 2 + (SALTO * ventas_ingresadas), column=columna7).value
                    ])
                if ws.cell(row=fila_producto + 3 + (SALTO * ventas_ingresadas), column=columna4).value is not None:
                    data.append([
                        ws.cell(row=fila_producto + 3 + (SALTO * ventas_ingresadas), column=columna4).value,
                        ws.cell(row=fila_producto + 3 + (SALTO * ventas_ingresadas), column=columna6).value,
                        ws.cell(row=fila_producto + 3 + (SALTO * ventas_ingresadas), column=columna7).value
                    ])

                if data:
                    print("PRODUCTOS INGRESADOS: ")
                    print(tabulate(data, headers=[], tablefmt="fancy_grid", stralign='center'))
                print("[recuerde incluir el envío en concepto producto]")
                productos_ingresados += 1
            else:
                productos_ingresados += 6

        coste_envio = int(input("Coste de Envío (Abonado por el Negocio): "))
        ws.cell(row=fila16 + (SALTO * ventas_ingresadas), column=columna2).value = coste_envio
        print("NO SE INGRESARON PAGOS DEL COMPRADOR AL PEDIDO")

        b = 0
        while b < A:
            pregunta = input("¿Ingresar Nuevo Pago del Comprador? SI/NO: ")
            if pregunta.upper() == "SI":
                monto_pago = int(input("Monto: "))
                ws.cell(row=fila17 + pagos_ingresados + (SALTO * ventas_ingresadas), column=columna7).value = monto_pago
                codigo_ext = input("Código de operación (plataformas): ")
                ws.cell(row=fila17 + pagos_ingresados + SALTO * ventas_ingresadas, column=columna6).value = codigo_ext
                codigo_medio = input("Cód. Medio de pago (ML/TR/MP/ECF/ESF): ")
                ws.cell(row=fila17 + pagos_ingresados + SALTO * ventas_ingresadas, column=columna5).value = codigo_medio

                data = []
                if ws.cell(row=fila17 + (SALTO * ventas_ingresadas), column=columna7).value is not None:
                    data.append([
                        ws.cell(row=fila17 + (SALTO * ventas_ingresadas), column=columna7).value,
                        ws.cell(row=fila17 + (SALTO * ventas_ingresadas), column=columna6).value,
                        ws.cell(row=fila17 + (SALTO * ventas_ingresadas), column=columna5).value
                    ])
                if ws.cell(row=fila17 + 1 + (SALTO * ventas_ingresadas), column=columna7).value is not None:
                    data.append([
                        ws.cell(row=fila17 + 1 + (SALTO * ventas_ingresadas), column=columna7).value,
                        ws.cell(row=fila17 + 1 + (SALTO * ventas_ingresadas), column=columna6).value,
                        ws.cell(row=fila17 + 1 + (SALTO * ventas_ingresadas), column=columna5).value
                    ])
                if ws.cell(row=fila17 + 2 + (SALTO * ventas_ingresadas), column=columna7).value is not None:
                    data.append([
                        ws.cell(row=fila17 + 2 + (SALTO * ventas_ingresadas), column=columna7).value,
                        ws.cell(row=fila17 + 2 + (SALTO * ventas_ingresadas), column=columna6).value,
                        ws.cell(row=fila17 + 2 + (SALTO * ventas_ingresadas), column=columna5).value
                    ])
                if ws.cell(row=fila17 + 3 + (SALTO * ventas_ingresadas), column=columna7).value is not None:
                    data.append([
                        ws.cell(row=fila17 + 3 + (SALTO * ventas_ingresadas), column=columna7).value,
                        ws.cell(row=fila17 + 3 + (SALTO * ventas_ingresadas), column=columna6).value,
                        ws.cell(row=fila17 + 3 + (SALTO * ventas_ingresadas), column=columna5).value
                    ])
                if ws.cell(row=fila17 + 4 + (SALTO * ventas_ingresadas), column=columna7).value is not None:
                    data.append([
                        ws.cell(row=fila17 + 4 + (SALTO * ventas_ingresadas), column=columna7).value,
                        ws.cell(row=fila17 + 4 + (SALTO * ventas_ingresadas), column=columna6).value,
                        ws.cell(row=fila17 + 4 + (SALTO * ventas_ingresadas), column=columna5).value
                    ])
                print("PAGOS INGRESADOS: ")
                print(tabulate(data, headers=[], tablefmt="fancy_grid", stralign='center'))

                pagos_ingresados += 1
            else:
                b += 1
        pregunta = input("¿Venta Facturada? SI/NO: ")
        ws.cell(row=fila23 + SALTO * ventas_ingresadas, column=columna2).value = pregunta
        wb.save("Libro1.xlsx")
    else:
        i += 1
